# 用对象方法读取excel表中的数据，并写一个返回的方法，让调用
from openpyxl import load_workbook   #导入openpyxl，用来读取excel表
from common.splicing import file_os1
from reports.request import Request
from reports.jm import decrypt_str
class Case:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.method=None
        self.url=None
        self.data=None
        self.headers=None
        self.expected=None
class Do_excel():
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def excel(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        case=[]
        for item in range(2,sheet.max_row+1):
            row=Case()
            row.case_id=sheet.cell(item,1).value
            row.title=sheet.cell(item,2).value
            row.method=sheet.cell(item,3).value
            row.url=sheet.cell(item,4).value
            row.data=sheet.cell(item,5).value
            row.headers = sheet.cell(item, 6).value
            row.expected=sheet.cell(item,7).value
            case.append(row)
        return case
    def write_data(self,row,actual,result):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        sheet.cell(row,8).value=actual
        sheet.cell(row, 9).value = result
        wb.save(self.file_name)
if __name__ == "__main__":
    do_excel=Do_excel(file_os1,"register")
    cases=do_excel.excel()

    for case in cases:
        print(case.data)

    #     request=Request()
    #     for case in cases:
    #         res=request.request(method=case.method,url=case.url,json=case.data)
    #         print(decrypt_str(res.json()))
    #     print(res.text)
        # if res.text==case.expected:
        #     do_excel.write_data(case.case_id+1,res.text,"pass")
        # else:
        #     do_excel.write_data(case.case_id+1,res.text,"Flas")


